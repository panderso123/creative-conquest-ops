from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY="1F3A5F"; GOLD="B08D57"; LIGHT="F2EEE6"; GREY="8A8A8A"; WHITE="FFFFFF"
BLUE="0000FF"; BLACK="000000"
thin=Side(style="thin", color="D9D2C5")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
bottom=Border(bottom=Side(style="thin",color=GOLD))
title_font=Font(name="Arial",size=18,bold=True,color=NAVY)
sub_font=Font(name="Arial",size=10,italic=True,color=GREY)
sec_font=Font(name="Arial",size=11,bold=True,color=WHITE)
sec_fill=PatternFill("solid",fgColor=NAVY)
lbl=Font(name="Arial",size=10,color="222222")
inp=Font(name="Arial",size=10,color=BLUE)          # hardcoded inputs = blue
calc=Font(name="Arial",size=10,color=BLACK,bold=True) # formulas = black
big=Font(name="Arial",size=14,bold=True,color=NAVY)
note=Font(name="Arial",size=8,italic=True,color=GREY)
ctr=Alignment(horizontal="center",vertical="center")
right=Alignment(horizontal="right")
band=PatternFill("solid",fgColor=LIGHT)

wb=Workbook(); ws=wb.active; ws.title="ROI Report"
ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=3
ws.column_dimensions["B"].width=34
ws.column_dimensions["C"].width=16
ws.column_dimensions["D"].width=16
ws.column_dimensions["E"].width=30

ws["B1"]="CREATIVE CONQUEST"; ws["B1"].font=Font(name="Arial",size=11,bold=True,color=GOLD)
ws["B2"]="Monthly ROI Report"; ws["B2"].font=title_font
ws["B3"]="The booked calendar and bought-back time - not 'video'. Blue = you fill in; black = auto-calculated."
ws["B3"].font=sub_font

# header block
hdr=[("Client",""),("Reporting month",""),("Tier / plan",""),("Monthly fee ($)",3250)]
r=5
for label,val in hdr:
    ws.cell(row=r,column=2,value=label).font=Font(name="Arial",size=10,bold=True,color=NAVY)
    cell=ws.cell(row=r,column=3,value=val)
    cell.font=inp if val!="" else inp
    cell.alignment=right
    if label=="Monthly fee ($)": cell.number_format='$#,##0'
    ws.cell(row=r,column=2).border=bottom; ws.cell(row=r,column=3).border=bottom
    r+=1

def section(row,text):
    ws.cell(row=row,column=2,value=text).font=sec_font
    for col in (2,3,4):
        ws.cell(row=row,column=col).fill=sec_fill
        ws.cell(row=row,column=col).border=border
    ws.cell(row=row,column=2).alignment=Alignment(horizontal="left",vertical="center")

def line(row,label,value,fmt=None,is_input=True,formula=False,note_txt=None):
    ws.cell(row=row,column=2,value=label).font=lbl
    cell=ws.cell(row=row,column=3,value=value)
    cell.font=inp if (is_input and not formula) else calc
    cell.alignment=right
    if fmt: cell.number_format=fmt
    ws.cell(row=row,column=2).border=border; ws.cell(row=row,column=3).border=border
    ws.cell(row=row,column=4).border=border
    if note_txt:
        ws.cell(row=row,column=5,value=note_txt).font=note
    return row+1

# ---- Reach & engagement ----
r=10
section(r,"1 · REACH & ENGAGEMENT"); r+=1
ws.cell(row=10,column=3,value="This month").font=Font(name="Arial",size=9,bold=True,color=WHITE)
ws.cell(row=10,column=3).alignment=ctr
videos_row=r
r=line(r,"Videos delivered to standard",20,'#,##0')
views_row=r
r=line(r,"Total views",0,'#,##0')
r=line(r,"Total engagements (likes+comments+shares+saves)",0,'#,##0')
eng_label_row=r
ws.cell(row=r,column=2,value="Engagement rate").font=lbl
er=ws.cell(row=r,column=3,value=f"=IF(C{views_row}=0,0,C{eng_label_row-1}/C{views_row})")
er.font=calc; er.number_format="0.0%"; er.alignment=right
ws.cell(row=r,column=2).border=border; ws.cell(row=r,column=3).border=border; ws.cell(row=r,column=4).border=border
ws.cell(row=r,column=5,value="Auto: engagements / views").font=note
r+=1
r=line(r,"Profile / website visits from content",0,'#,##0')
r=line(r,"Follower growth (net new)",0,'#,##0')

# ---- Pipeline ----
r+=1
section(r,"2 · PIPELINE & BOOKINGS"); r+=1
dm_row=r
r=line(r,"Inbound DMs / inquiries",0,'#,##0')
consult_row=r
r=line(r,"Consults booked",0,'#,##0')
r=line(r,"Consults attributed to content",0,'#,##0')
newclient_row=r
r=line(r,"New clients closed from content",0,'#,##0')

# ---- Revenue & ROI ----
r+=1
section(r,"3 · REVENUE & ROI"); r+=1
val_row=r
r=line(r,"Avg. annual value per new client ($)",2250,'$#,##0',note_txt="Playbook: $1,500-$3,000+/yr")
rev_row=r
ws.cell(row=r,column=2,value="Est. revenue attributed (annualized)").font=Font(name="Arial",size=10,bold=True,color=NAVY)
rv=ws.cell(row=r,column=3,value=f"=C{newclient_row}*C{val_row}")
rv.font=calc; rv.number_format='$#,##0'; rv.alignment=right
ws.cell(row=r,column=2).border=border; ws.cell(row=r,column=3).border=border; ws.cell(row=r,column=4).border=border
ws.cell(row=r,column=5,value="Auto: new clients x annual value").font=note
r+=1
fee_row=r
r=line(r,"Monthly fee ($)",f"=C8",'$#,##0',formula=True)
roi_row=r
ws.cell(row=r,column=2,value="ROI multiple (annual rev / monthly fee)").font=Font(name="Arial",size=10,bold=True,color=NAVY)
roi=ws.cell(row=r,column=3,value=f"=IF(C{fee_row}=0,0,C{rev_row}/C{fee_row})")
roi.font=Font(name="Arial",size=12,bold=True,color=GOLD); roi.number_format='0.0"x"'; roi.alignment=right
ws.cell(row=r,column=2).border=border; ws.cell(row=r,column=3).border=border; ws.cell(row=r,column=4).border=border
r+=1
ws.cell(row=r,column=2,value="Cost per consult booked ($)").font=lbl
cpc=ws.cell(row=r,column=3,value=f"=IF(C{consult_row}=0,0,C{fee_row}/C{consult_row})")
cpc.font=calc; cpc.number_format='$#,##0'; cpc.alignment=right
ws.cell(row=r,column=2).border=border; ws.cell(row=r,column=3).border=border; ws.cell(row=r,column=4).border=border
r+=2

# Highlight callout
ws.cell(row=r,column=2,value="THE HEADLINE").font=Font(name="Arial",size=10,bold=True,color=GOLD)
r+=1
ws.cell(row=r,column=2,value=f'=CONCATENATE("Delivered ",C{videos_row}," videos. ",TEXT(C{views_row},"#,##0")," views. ",C{consult_row}," consults booked.")')
ws.cell(row=r,column=2).font=big
r+=2
ws.cell(row=r,column=2,value="Notes / wins this month:").font=Font(name="Arial",size=10,bold=True,color=NAVY)
r+=1
for _ in range(3):
    ws.cell(row=r,column=2).border=bottom
    ws.cell(row=r,column=3).border=bottom
    ws.cell(row=r,column=4).border=bottom
    ws.cell(row=r,column=5).border=bottom
    r+=1

# Sheet 2: instructions
g=wb.create_sheet("How to use")
g.sheet_view.showGridLines=False
g.column_dimensions["A"].width=3; g.column_dimensions["B"].width=90
g["B1"]="How to use this report"; g["B1"].font=title_font
tips=[
 "Fill only the BLUE cells each month - everything in black calculates automatically.",
 "Pull reach numbers from native analytics (Instagram/TikTok/YouTube insights) or Buffer/Meta Suite.",
 "'Attributed to content' = bookings where the lead came from a Reel/Short/TikTok or a content-driven DM. Ask at intake: 'How did you hear about us?'",
 "Avg. annual client value: confirm with the client at onboarding; playbook range is $1,500-$3,000+/yr. Default is $2,250.",
 "ROI multiple is the number that renews the retainer - lead the monthly call with it.",
 "Growth tier = monthly report + call. Authority tier = bi-weekly. Duplicate this file per client per month.",
]
rr=3
for t in tips:
    g.cell(row=rr,column=2,value="- "+t).font=Font(name="Arial",size=10,color="222222")
    g.cell(row=rr,column=2).alignment=Alignment(wrap_text=True,vertical="top")
    g.row_dimensions[rr].height=30
    rr+=1
g.cell(row=rr+1,column=2,value="Guarantee reminder: satisfaction-based on quality of work - never promise specific income or booking numbers.").font=Font(name="Arial",size=9,italic=True,color=GREY)

wb.save("Monthly_ROI_Report.xlsx")
print("saved roi")

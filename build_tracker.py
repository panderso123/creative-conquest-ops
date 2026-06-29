from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

NAVY="1F3A5F"; GOLD="B08D57"; LIGHT="F2EEE6"; GREY="8A8A8A"; WHITE="FFFFFF"
thin=Side(style="thin", color="D9D2C5")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
title_font=Font(name="Arial", size=16, bold=True, color=NAVY)
sub_font=Font(name="Arial", size=10, italic=True, color=GREY)
hdr_font=Font(name="Arial", size=10, bold=True, color=WHITE)
cell_font=Font(name="Arial", size=10, color="222222")
hdr_fill=PatternFill("solid", fgColor=NAVY)
band_fill=PatternFill("solid", fgColor=LIGHT)
wrap=Alignment(wrap_text=True, vertical="top")
ctr=Alignment(horizontal="center", vertical="center", wrap_text=True)

wb=Workbook()
ws=wb.active; ws.title="Production Tracker"
ws.sheet_view.showGridLines=False
ws["A1"]="CREATIVE CONQUEST  -  Content Production Tracker"; ws["A1"].font=title_font
ws["A2"]="20 videos / month to standard - film -> edit -> QA -> approve -> schedule. One row per video."
ws["A2"].font=sub_font
ws["A3"]="Client:"; ws["B3"]="________________"
ws["D3"]="Month:"; ws["E3"]="________________"
ws["G3"]="Tier:"; ws["H3"]="Growth (20/mo)"
for c in ["A3","D3","G3"]: ws[c].font=Font(name="Arial",size=10,bold=True,color=NAVY)
for c in ["B3","E3","H3"]: ws[c].font=cell_font

headers=["#","Working Title","Hook / Angle","Content Pillar","Film Date",
         "Editor","Edit Status","QA","Client Approved","Schedule Date",
         "Platform(s)","Posted?","Post Link","Notes"]
widths=[4,24,28,16,11,12,14,9,15,12,16,9,22,26]
hrow=5
for i,h in enumerate(headers,1):
    c=ws.cell(row=hrow,column=i,value=h); c.font=hdr_font; c.fill=hdr_fill
    c.alignment=ctr; c.border=border
    ws.column_dimensions[chr(64+i)].width=widths[i-1]

pillars=["Education","Myth-Buster","Treatment Demo","Patient Result","Behind the Scenes",
         "FAQ / Objection","Provider Spotlight","Trend / Seasonal"]
for r in range(hrow+1, hrow+21):
    n=r-hrow
    ws.cell(row=r,column=1,value=n)
    for col in range(1,15):
        cell=ws.cell(row=r,column=col)
        cell.border=border; cell.font=cell_font; cell.alignment=wrap
        if n%2==0: cell.fill=band_fill

dv_edit=DataValidation(type="list",formula1='"Not Started,Filming,Editing,Revisions,Done"',allow_blank=True)
dv_qa=DataValidation(type="list",formula1='"Pending,Pass,Redo"',allow_blank=True)
dv_app=DataValidation(type="list",formula1='"Awaiting,Approved,Changes Requested"',allow_blank=True)
dv_post=DataValidation(type="list",formula1='"No,Scheduled,Posted"',allow_blank=True)
dv_plat=DataValidation(type="list",formula1='"Reels,TikTok,Shorts,Reels+TikTok,All 3,All + LinkedIn"',allow_blank=True)
dv_pillar=DataValidation(type="list",formula1='"'+",".join(pillars)+'"',allow_blank=True)
for dv in (dv_edit,dv_qa,dv_app,dv_post,dv_plat,dv_pillar): ws.add_data_validation(dv)
dv_pillar.add(f"D{hrow+1}:D{hrow+20}")
dv_edit.add(f"G{hrow+1}:G{hrow+20}")
dv_qa.add(f"H{hrow+1}:H{hrow+20}")
dv_app.add(f"I{hrow+1}:I{hrow+20}")
dv_plat.add(f"K{hrow+1}:K{hrow+20}")
dv_post.add(f"L{hrow+1}:L{hrow+20}")
ws.freeze_panes="A6"

ps=wb.create_sheet("Pipeline Summary")
ps.sheet_view.showGridLines=False
ps["A1"]="Pipeline Summary"; ps["A1"].font=title_font
ps["A2"]="Live counts from the Production Tracker. Target = 20 delivered to standard each month."
ps["A2"].font=sub_font
rows=[
 ("Videos in tracker","=COUNTA('Production Tracker'!B6:B25)"),
 ("Editing in progress",'=COUNTIF(\'Production Tracker\'!G6:G25,"Editing")'),
 ("In revisions",'=COUNTIF(\'Production Tracker\'!G6:G25,"Revisions")'),
 ("Edits done",'=COUNTIF(\'Production Tracker\'!G6:G25,"Done")'),
 ("Passed QA",'=COUNTIF(\'Production Tracker\'!H6:H25,"Pass")'),
 ("Client approved",'=COUNTIF(\'Production Tracker\'!I6:I25,"Approved")'),
 ("Scheduled or posted",'=COUNTIF(\'Production Tracker\'!L6:L25,"Scheduled")+COUNTIF(\'Production Tracker\'!L6:L25,"Posted")'),
 ("Posted (live)",'=COUNTIF(\'Production Tracker\'!L6:L25,"Posted")'),
]
ps["A4"]="Metric"; ps["B4"]="Count"
for c in ("A4","B4"): ps[c].font=hdr_font; ps[c].fill=hdr_fill; ps[c].border=border; ps[c].alignment=ctr
r=5
for label,formula in rows:
    ps.cell(row=r,column=1,value=label).font=cell_font
    ps.cell(row=r,column=2,value=formula).font=cell_font
    ps.cell(row=r,column=1).border=border; ps.cell(row=r,column=2).border=border
    ps.cell(row=r,column=2).alignment=ctr
    r+=1
posted_row=12
target_row=r+1
ps.cell(row=target_row,column=1,value="Monthly target").font=Font(name="Arial",size=10,bold=True,color=NAVY)
ps.cell(row=target_row,column=2,value=20).font=Font(name="Arial",size=10,bold=True,color="0000FF")
ps.cell(row=target_row,column=2).alignment=ctr
ps.cell(row=target_row+1,column=1,value="% to target (posted)").font=Font(name="Arial",size=10,bold=True,color=NAVY)
pc=ps.cell(row=target_row+1,column=2,value=f"=B{posted_row}/B{target_row}")
pc.font=Font(name="Arial",size=10,bold=True,color=NAVY); pc.number_format="0%"; pc.alignment=ctr
ps.column_dimensions["A"].width=26; ps.column_dimensions["B"].width=12

cal=wb.create_sheet("Monthly Calendar")
cal.sheet_view.showGridLines=False
cal["A1"]="Posting Calendar - schedule ~5/week"; cal["A1"].font=title_font
cal["A2"]="Drop each video's working title into a day once it clears QA and approval."
cal["A2"].font=sub_font
days=["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
for i,d in enumerate(days,1):
    c=cal.cell(row=4,column=i,value=d); c.font=hdr_font; c.fill=hdr_fill; c.alignment=ctr; c.border=border
    cal.column_dimensions[chr(64+i)].width=18
for wk in range(4):
    base=5+wk*3
    cal.cell(row=base,column=9,value=f"Week {wk+1}").font=Font(name="Arial",size=9,bold=True,color=GREY)
    for i in range(1,8):
        for rr in range(base,base+2):
            cell=cal.cell(row=rr,column=i); cell.border=border; cell.font=cell_font; cell.alignment=wrap
            if wk%2==0: cell.fill=band_fill
    cal.row_dimensions[base].height=22; cal.row_dimensions[base+1].height=22

wb.save("Content_Production_Tracker.xlsx")
print("saved tracker")
